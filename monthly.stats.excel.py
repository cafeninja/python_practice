# from openpyxl import load_workbook
#
# workbook = load_workbook('schedule_test.xlsx', read_only=True)
# second_sheet = 'April'
# #worksheet = workbook.get_sheet_by_name(second_sheet)
# worksheet = wb[second_sheet]
#
# for row in worksheet.iter_rows():
#     print (row)
#
# # check out the last row
# for cell in row:
#     print (cell)
#
import re
import string
import datetime
import dateutil.relativedelta
import contextlib
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Find OT file.  If there, delete... then create a tnew one.202004_NOC OT_reporting.xlsx
now = datetime.now()
filename = 'Test.Ticket.Data.Working.File.xlsx'
#print(filename)

## count total count of incidents

## count detect = monitoring

## count government tickets

## count provider tickets

## count commercial tickets

## count platform tickets


## count # of government = detected by monitoring
## count # of provider = detected by monitoring
## count # of commercial = detected by monitoring
## count # of platform = detected by monitoring

## 
##  Make xlxs and input each row into it.
