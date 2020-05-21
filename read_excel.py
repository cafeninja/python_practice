# from openpyxl import load_workbook
#
# workbook = load_workbook('schedule_test.xlsx', read_only=True)
# second_sheet = 'April'
# #worksheet = workbook.get_sheet_by_name(second_sheet)
# worksheet = wb[second_sheet]
#
# for row in worksheet.iter_rows():
#     print (row)
#
# # check out the last row
# for cell in row:
#     print (cell)
#
import re
import string
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('april.schedule.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

# Team members by employee number:
staffdict = {
"Terence Briffa (tebr01)": 9865,
"Christian Scicluna (chsc01)": 3985,
"Nicholai Mifsud (nimi)": 2879,
"Dennis Giordmaina (degi)": 3098,
"Luca Distinto (ludi)": 2975,
"Godwin Urio (gour01)": 10872,
"Jonathan Mizzi (jomi)": 2995,
"Kenrick Bartolo (keba01)": 11248,
"Jurgen Sammut (jusa01)": 11495,
"Jean Claude Aber (jeab02)": 11452,
"Christopher Muscat (chmu03)": 11505,
"Andrea Gaffiero (anga02)": 11504
}
#print(staffdict)  # for troublehsooting

print('------------')
for row in worksheet.iter_rows():
    for cell in row:
        if cell.comment:
            fmtrow = []
            name = str(row[0].value)
            namelist = name.split(" ")
            #print(row[0].value)
            cellname = str(cell)
            commentRaw = str(cell.comment.text)
            col2 = str(re.findall("\.[A-Z]{1,2}", cellname))
            col1 = str(re.findall("[A-Z]+", col2))
            col1 = re.sub('[\'\[\]!@#$]', '', col1)
            row2 = str(re.findall("[A-Z]\d+", cellname))
            row1 = re.sub('[A-Z]+', '', row2)
            row1 = re.sub('[\'\[\]!@#$]', '', row1)
            rownum = re.findall("[0-9]{1,2}", cellname)
            date = col1+'2'
            cal = str((worksheet[date].value))+ ' ' + str((worksheet["B1"].value))
            # ==============================================
            # print(cellname)  # for troubleshooting
            #print(f'Column letter: {col1}')  # for troubleshooting
            #print(f'Row number: {rownum}...{row1}')  # for troubleshooting
            #print(f'Date cell: {date}')  # for troubleshooting            
            # ==============================================
            # GET EMPLOYEE NUMBER, FIRST NAME and LAST NAME OR PRINT HOLIDAY
            if len(namelist) > 1:
                fmtrow.append(staffdict.get(name))
                # print(staffdict.get(name))
                fmtrow.append(namelist[0])
                #print(namelist[0])
                fmtrow.append(namelist[1])
                #print(namelist[1])
            elif len(namelist) <= 1:
                note = str(re.sub('[\n\'\[\]!@#$]', '', commentRaw))
                fmtrow.append(note)
                #print(note)
            # Add Date to list for output.
            fmtrow.append(cal)
            #print(cal)
            # Discover if date is a sunday.  USE OT15, Unless Sundeay, then OT20
            dayweek = datetime.strptime(cal, '%d %B %Y')
            # print(dayweek)  # for troubleshooting
            if dayweek.weekday() == 6:
                fmtrow.append('OT20')
                #print(f'OT20')
            elif dayweek.weekday() != 6:
                fmtrow.append('OT15')
                #print(f'OT15')
            #Get Hours and Reason from comment
            if re.findall("-", commentRaw):
                ottime = commentRaw.split('-')
                hrs = str(re.findall("\d+", ottime[0]))
                hrs = re.sub('[\'\[\]!@#$]', '', hrs)
                fmtrow.append(hrs)
                #print(hrs)
                fmtrow.append(ottime[1])
                #print(ottime[1])
            # show collected output whic will become the feed for new XLXS
            print(fmtrow, sep=" | ")

# Format of output needed for HR.
# Employee Code	| Name of Employee  | Surname   | Date          | Overtime Code	| Hours
# 123	        | Nadine	    | Dalli	| 2/8/2017	| OT15	        | 3

print('------------')
