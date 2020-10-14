# from openpyxl import load_workbook
#
# workbook = load_workbook('schedule_test.xlsx', read_only=True)
# second_sheet = 'April'
# #worksheet = workbook.get_sheet_by_name(second_sheet)
# worksheet = wb[second_sheet]
#
# for row in worksheet.iter_rows():
#     print (row)
#
# # check out the last row
# for cell in row:
#     print (cell)
#
import re
import string
import datetime
import dateutil.relativedelta
import contextlib
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Find OT file.  If there, delete... then create a tnew one.202004_NOC OT_reporting.xlsx
now = datetime.now()
fnlastmonth = (now + dateutil.relativedelta.relativedelta(months=-1)).strftime("%Y%m")
filename = fnlastmonth + '_NOC_OT_reporting.xlsx'
#print(filename)

with contextlib.suppress(FileNotFoundError):
    os.remove(filename)
book = Workbook()
sheet = book.active
titlerow = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
sheet['A1'] = 'Employee Code'
sheet['B1'] = 'Name of Employee'
sheet['C1'] = 'Surname of Employee'
sheet['D1'] = 'Date of Overtime'
sheet['E1'] = 'Overtime Code'
sheet['F1'] = 'Hours'
sheet['G1'] = 'Reason'
for cell in titlerow:
    sheet[cell].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet[cell].font = Font(bold=True)
for column in columns:
    sheet.column_dimensions[column].width = 22
book.save(filename)


workbook = load_workbook('2020_Malta_NOC_Schedule.xlsx')
first_sheet = workbook.sheetnames[1]
print(f'Worksheet name: {first_sheet}')
worksheet = workbook[first_sheet]
#worksheet = workbook.get_sheet_by_name(first_sheet)

# Team members by employee number:
staffdict = {
"Terence Briffa (tebr01)": 9865,
"Christian Scicluna (chsc01)": 3985,
"Nicholai Mifsud (nimi)": 2872,
"Dennis Giordmaina (degi)": 3098,
"Luca Distinto (ludi)": 2975,
"Godwin Urio (gour01)": 10872,
"Jonathan Mizzi (jomi)": 2995,
"Kenrick Bartolo (keba01)": 11248,
"Jurgen Sammut (jusa01)": 11495,
"Jean Claude Aber (jeab02)": 11452,
"Christopher Muscat (chmu03)": 11505,
"Andrea Gaffiero (anga02)": 11504
}
#print(staffdict)  # for troublehsooting
holidaylist = []
print('------------')
for row in worksheet.iter_rows():
    for cell in row:
        if cell.comment:
            fmtrow = []
            name = str(row[0].value)
            namelist = name.split(" ")
            #print(row[0].value)
            cellname = str(cell)
            commentRaw = str(cell.comment.text)
            commentRaw = str(re.sub(r"^.*\n", '', commentRaw))
            commentRaw = str(re.sub(r"^.*\n", '', commentRaw))
            commentRaw = str(re.sub(r"^.*\n", '', commentRaw))
            commentRaw = str(re.sub(r"^.*\n", '', commentRaw))
            commentRaw = commentRaw.strip()
            col2 = str(re.findall("\.[A-Z]{1,2}", cellname))
            col1 = str(re.findall("[A-Z]+", col2))
            col1 = re.sub('[\'\[\]!@#$]', '', col1)
            row2 = str(re.findall("[A-Z]\d+", cellname))
            row1 = re.sub('[A-Z]+', '', row2)
            row1 = re.sub('[\'\[\]!@#$]', '', row1)
            rownum = re.findall("[0-9]{1,2}", cellname)
            date = col1+'2'
            mon = col1+'1'
            # print(mon) # for troubleshooting
            cal = str((worksheet[date].value))+ ' ' + str((worksheet["B1"].value))
            # ==============================================
            # print(cellname)  # for troubleshooting
            #print(f'Column letter: {col1}')  # for troubleshooting
            #print(f'Row number: {rownum}...{row1}')  # for troubleshooting
            #print(f'Date cell: {date}')  # for troubleshooting            
            # ==============================================
            # GET EMPLOYEE NUMBER, FIRST NAME and LAST NAME OR PRINT HOLIDAY
            if len(namelist) > 1:
                fmtrow.append(staffdict.get(name))
                # print(staffdict.get(name))
                fmtrow.append(namelist[0])
                #print(namelist[0])
                fmtrow.append(namelist[1])
                #print(namelist[1])
            elif len(namelist) <= 1:
                note = str(re.sub('[\n\'\[\]!@#$]', '', commentRaw))
                #note = "Identified Holiday: " + note
                fmtrow.append("Identified Holiday:")
                fmtrow.append(note)
                fmtrow.append('.')
                holidaylist.append(cell.column)
                #print(note)
            # Add Date to list for output.
            fmtrow.append(cal)
            #print(cal)
            # Discover if date is a sunday.  USE OT15, Unless Sundeay, then OT20
            dayweek = datetime.strptime(cal, '%d %B %Y')
            # print(dayweek)  # for troubleshooting
            if dayweek.weekday() == 6:
                fmtrow.append('OT20')
                #print(f'OT20')
            elif dayweek.weekday() != 6:
                fmtrow.append('OT15')
                #print(f'OT15')
            #Get Hours and Reason from comment
            if re.findall("-", commentRaw):
                ottime = commentRaw.split('-')
                hrs = str(re.findall("\d+", ottime[0]))
                hrs = re.sub('[\'\[\]!@#$]', '', hrs)
                fmtrow.append(hrs)
                #print(hrs)
                fmtrow.append(ottime[1])
                #print(ottime[1])
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
            ws.append(fmtrow)
            wb.save(filename)
#========================
            print(fmtrow, sep=" | ")

# Format of output needed for HR.
# Employee Code	| Name of Employee  | Surname   | Date          | Overtime Code	| Hours
# 123	        | Nadine	    | Dalli	| 2/8/2017	| OT15	        | 3
hebreak = ['----Holidy Entries--------', '----Holidy Entries--------']
wb = load_workbook(filename)
ws = wb.worksheets[0]
ws.append(hebreak)
wb.save(filename)
print('----Holidy Entries--------')
#print(holidaylist)   # for troubleshooting
for i in holidaylist:
    # print(i)   # for troubleshooting
    #holidaycol = str(i+"3:"+i+"15")
    # print(holidaycol)
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.column == i:
                if cell.value == 'D' or cell.value == 'N' or cell.value == '2N' or cell.value == '2D':
                    fmtrow = []
                    name = str(row[0].value)
                    namelist = name.split(" ")
                    #print(row[0].value)
                    cellname = str(cell)
                    #commentRaw = str(cell.comment.text)
                    col2 = str(re.findall("\.[A-Z]{1,2}", cellname))
                    col1 = str(re.findall("[A-Z]+", col2))
                    col1 = re.sub('[\'\[\]!@#$]', '', col1)
                    row2 = str(re.findall("[A-Z]\d+", cellname))
                    row1 = re.sub('[A-Z]+', '', row2)
                    row1 = re.sub('[\'\[\]!@#$]', '', row1)
                    rownum = re.findall("[0-9]{1,2}", cellname)
                    date = col1+'2'
                    cal = str((worksheet[date].value))+ ' ' + str((worksheet["B1"].value))
                    # ==============================================
                    # print(cellname)  # for troubleshooting
                    #print(f'Column letter: {col1}')  # for troubleshooting
                    #print(f'Row number: {rownum}...{row1}')  # for troubleshooting
                    #print(f'Date cell: {date}')  # for troubleshooting            
                    # ==============================================
                    # GET EMPLOYEE NUMBER, FIRST NAME and LAST NAME OR PRINT HOLIDAY
                    if len(namelist) > 1:
                        fmtrow.append(staffdict.get(name))
                        # print(staffdict.get(name))
                        fmtrow.append(namelist[0])
                        #print(namelist[0])
                        fmtrow.append(namelist[1])
                        #print(namelist[1])
                    elif len(namelist) <= 1:
                        note = "Fail" # old content :: str(re.sub('[\n\'\[\]!@#$]', '', commentRaw))
                        fmtrow.append(note)
                        holidaylist.append(col1)
                        #print(note)
                    # Add Date to list for output.
                    fmtrow.append(cal)
                    fmtrow.append('OT20')
                    fmtrow.append('12')
                    wb = load_workbook(filename)
                    ws = wb.worksheets[0]
                    ws.append(fmtrow)
                    wb.save(filename)
                    print(fmtrow, sep=" | ")

## Add list of persons working on feast which is not on Saturday or Sunday
sbreak = ['----Scheduled off Weekday Feast--------', '----Scheduled off Weekday Feast--------']
wb = load_workbook(filename)
ws = wb.worksheets[0]
ws.append(sbreak)
wb.save(filename)
print('----Scheduled off Weekday Feast--------')
#print(holidaylist)   # for troubleshooting
for i in holidaylist:
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.column == i:
                if cell.value is None:
                    fmtrow = list(range(4))
                    name = str(row[0].value)
                    namelist = name.split(" ")
                    #print(row[0].value)
                    cellname = str(cell)
                    #commentRaw = str(cell.comment.text)
                    col2 = str(re.findall("\.[A-Z]{1,2}", cellname))
                    col1 = str(re.findall("[A-Z]+", col2))
                    col1 = re.sub('[\'\[\]!@#$]', '', col1)
                    row2 = str(re.findall("[A-Z]\d+", cellname))
                    row1 = re.sub('[A-Z]+', '', row2)
                    row1 = re.sub('[\'\[\]!@#$]', '', row1)
                    rownum = re.findall("[0-9]{1,2}", cellname)
                    date = col1+'2'
                    cal = str((worksheet[date].value))+ ' ' + str((worksheet["B1"].value))
                    dayweek = datetime.strptime(cal, '%d %B %Y')
                    # print(dayweek)  # for troubleshooting
                    if dayweek.weekday() == 6 or dayweek.weekday() == 5:
                        del fmtrow[3]
                        #@print(f'Weekend')
                    elif dayweek.weekday() < 6:
                        fmtrow[3] = 'Day Off Needed'
                    # ==============================================
                    # print(cellname)  # for troubleshooting
                    #print(f'Column letter: {col1}')  # for troubleshooting
                    #print(f'Row number: {rownum}...{row1}')  # for troubleshooting
                    #print(f'Date cell: {date}')  # for troubleshooting            
                    # ==============================================
                    # GET EMPLOYEE NUMBER, FIRST NAME and LAST NAME OR PRINT HOLIDAY
                    if len(namelist) > 1:
                        #fmtrow.append(staffdict.get(name))
                        # print(staffdict.get(name))
                        fmtrow[0] = namelist[0]
                        #print(namelist[0])
                        fmtrow[1] = namelist[1]
                        #print(namelist[1])
                    elif len(namelist) <= 1:
                        note = "Fail" # old content :: str(re.sub('[\n\'\[\]!@#$]', '', commentRaw))
                        fmtrow.append(note)
                        holidaylist.append(col1)
                        #print(note)
                    # Add Date to list for output.
                    fmtrow[2] = cal
                    if len(fmtrow) == 4 and len(fmtrow[0]) > 2:
                        wb = load_workbook(filename)
                        ws = wb.worksheets[0]
                        ws.append(fmtrow)
                        wb.save(filename)
                        print(fmtrow, sep=" | ")
##  Make xlxs and input each row into it.
